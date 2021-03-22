import openpyxl as xl 
import os
InputListPath = []
keysearch=[]
select = 0

masterpath = "D:\\MasterSheet.xlsx"

#----Function to check masterbook active sheet--------#
def checkstatus():
    loadmaster = xl.load_workbook(masterpath)
    mastersheet = loadmaster['Sheet1']
    # mastermaxcol = mastersheet.max_column
    # mastermaxrow = mastersheet.max_row    
    mastersheet.delete_rows(2, mastersheet.max_row - 1)
    loadmaster.save(masterpath)
currrow = 0
currcol = 0

#----Function to Print header row--------#
def print_header(lpath):
    loaded_workbook1 = xl.load_workbook(lpath)
    numsheet = loaded_workbook1.sheetnames
    loadingsheet = loaded_workbook1[numsheet[0]]
    loadmaster = xl.load_workbook(masterpath)
    mastersheet = loadmaster['Sheet1']
    max_row_master = mastersheet.max_row
    max_row_master = max_row_master+3
    max_col_sheet1 = loadingsheet.max_column
    for c in range(1, max_col_sheet1+1):     # changed to max_col_sheet1 in lace 4
        mastersheet.cell(row = max_row_master, column= c).value = loadingsheet.cell(row= 1, column= c).value
    currcol = mastersheet.max_column
    # mastersheet.font  = Font(color="FF0000")
    loadmaster.save(masterpath)
    return currcol

#----Function to search and print data in mastersheet--------#
def searchandprint(path):
    lpath = path
    loaded_workbook1 = xl.load_workbook(path)
    numsheet = loaded_workbook1.sheetnames
    # loadingsheet = loaded_workbook1[numsheet[0]]
    # colloadingsheet = loadingsheet.max_column
    # print(numsheet)
    print("Data Write Successfully : Check The Master File")
    lensheet = len(numsheet)
    
    loadmaster = xl.load_workbook(masterpath)
    mastersheet = loadmaster['Sheet1']

    colcurr = print_header(lpath)
    # savecol = colcurr
    loadmaster = xl.load_workbook(masterpath)
    mastersheet = loadmaster['Sheet1']
    a = 1
    # mastercol = 0
#----Function to search data by PS Number--------#    
    for i in range(0, lensheet):
        activesheet = loaded_workbook1[numsheet[i]]
        # mastermaxcol = mastersheet.max_column
        mastermaxrow = mastersheet.max_row
        # temp_c = mastermaxcol+1
        maxrows = activesheet.max_row
        maxcol = activesheet.max_column
        # print("Max Rows and Coloumn are :",maxrows,maxcol)
        for rows in range(2, maxrows+1):
            if select == 1: 
                for col in range(1,3):
                    cellvalue1 = activesheet.cell(row= rows, column= col)
                    if str(cellvalue1.value) == str(keysearch[0]):         
                        if a == 1:   # printing name once
                            mastermaxrow = mastersheet.max_row
                            mastermaxrow = mastermaxrow+1
                            for r in range(1, 4):
                               mastersheet.cell(row= mastermaxrow, column = r).value = activesheet.cell(row= rows, column= r).value 
                               r = r+1
                            a = a+1
                        mastermaxrow = mastersheet.max_row +1
                        k = 4
                        for temp in range(4, maxcol+1):
                            mastersheet.cell(row= mastermaxrow, column = k).value = activesheet.cell(row= rows, column= temp).value
                            k = k+1
#----Function to search data by Name--------#                             
            elif select == 0: 
                for col in range(1,3):
                    cellvalue1 = activesheet.cell(row= rows, column= col)
                    if str(cellvalue1.value) == str(keysearch[0]): 
                        
                        if a == 1:   # printing name once
                            mastermaxrow = mastersheet.max_row
                            mastermaxrow = mastermaxrow+1
                            for r in range(1, 4):
                               mastersheet.cell(row= mastermaxrow, column = r).value = activesheet.cell(row= rows, column= r).value 
                               r = r+1
                            a = a+1
                        mastermaxrow = mastersheet.max_row
                        k = 4
                        for temp in range(4, maxcol+1):
                            mastersheet.cell(row= mastermaxrow, column = k).value = activesheet.cell(row= rows, column= temp).value
                            k = k+1   
    loadmaster.save(str('MasterSheet.xlsx'))
#----Function to take user input choice for file path --------#    
def userpathinput():
    path = input("Enter your Woorkbook Path : ")
    InputListPath.append(path)
    while True:
        choice = input("Do You Want To Add More Workbook :: Y/N: ")
        if choice == "Y" or choice == "y":
            path = input("Paste new path here: ")
            InputListPath.append(path)
        elif choice == "N" or choice == "n":
            break
#----Function to take user input choice for Input Choice --------#            
def Inputsearchkey():
    d1a = input (" Do you want to: A)\033[1m Search By PS Number :\033[0m B) \033[1mSearch By User Name :\033[0m [A/B]? :  ")
    if d1a == "A" or d1a == "a":
        user_PS_number = int(input("Enter PS number : "))
        keysearch.append(user_PS_number)  
        return 1 
    if d1a == "B" or d1a == "b": 
        user_name = str(input("Enter user name : "))
        keysearch.append(user_name)
        return 0
userpathinput()
select = Inputsearchkey()
lenpathlist = len(InputListPath)
checkstatus()
for i in range(0,lenpathlist):
    searchandprint(InputListPath[i])
